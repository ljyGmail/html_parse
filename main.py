import re
import os
from openpyxl import Workbook

# XPATH 정규 표현식
xpath_regex = re.compile(r'Table-Group aClass=([^<★]*)')
# 컬럼명 정규 표현식
item_regex = re.compile(r'u★([^\s]*) \|List\|')

target_file_list = [
'00634', '00636', '00637', '10001', '10002', '10003', '10081', '10084', '10085', '11011', '11012', '11013', '11200', '11301', '11303', '11304', '11305', '11306', '11307', '11308',
'11309', '11310', '11312', '11313', '11314', '11315', '11316', '11317', '11324', '11325', '11326', '11327', '11329', '11332', '11333', '11334', '11335', '11336', '11337', '11338',
'11339', '11340', '11341', '11342', '11343', '11344', '11345', '11346', '11347'
]

result = []
for file in sorted(os.listdir('html_files')):
    if target_file_list.count(file[0:5]):
        print(file)

        class_list = []

        with open('html_files/' + file, encoding='utf-8') as f:
            print('======================')
            content = f.read()
            class_mo = xpath_regex.search(content)  # XPATH 값 찾기

            start_end_list = []  # XPATH 의 시작 index 과 끝 index 저장

            while class_mo is not None:
                print(class_mo.group(1))
                class_list.append(class_mo.group(1))
                start_end_list.append((class_mo.start(), class_mo.end()))
                print('@' * 30)
                class_mo = xpath_regex.search(content, class_mo.end())

        print(start_end_list)
        print(class_list)

        flat_list = []
        for idx_tpl in start_end_list:
            flat_list.append(idx_tpl[0])
            flat_list.append(idx_tpl[1])

        print(flat_list)

        class_idx_list = []

        for i in range(len(start_end_list)):
            if len(flat_list) != 2 * i + 2:
                group_tuple = (flat_list[2 * i + 1], flat_list[2 * i + 2])
            else:
                group_tuple = (flat_list[2 * i + 1], len(content))

            class_idx_list.append(group_tuple)

        print(class_idx_list)

        result_class_list = []

        for idx, class_idx in enumerate(class_idx_list):
            class_content = content[class_idx[0]:class_idx[1]]
            print(class_content)
            print('##########')
            item_mo = item_regex.search(class_content)

            item_list = []
            while item_mo is not None:
                item_nm = item_mo.group(1)
                item_list.append(item_nm)
                print(item_nm)
                item_mo = item_regex.search(class_content, item_mo.end())
                print('************')

            if len(item_list) > 0:
                class_dict = {class_list[idx]: item_list}
                result_class_list.append(class_dict)

        print(result_class_list)

        result_dict = {file: result_class_list}
        result.append(result_dict)

print(result)

wb = Workbook()  # 새 워크북 생성

ws = wb.active  # 현재 활성화된 sheet 가져옴
ws.title = 'HTML_PARSE'  # sheet의 이름을 변경

count = 1

for f in result:
    file_name = list(f.keys())[0]
    file_class_list = list(f.values())[0]
    for file_class in file_class_list:
        class_name = list(file_class.keys())[0]
        for column_nm in list(file_class.values())[0]:
            print(f'{file_name},{class_name},{column_nm}')
            ws.cell(row=count, column=1, value=file_name[0:5])
            ws.cell(row=count, column=2, value=class_name)
            ws.cell(row=count, column=3, value=column_nm)
            count += 1

wb.save('HTML_parsing_result.xlsx')
wb.close()
